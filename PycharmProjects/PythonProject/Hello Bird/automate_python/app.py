import openpyxl as xl
from openpyxl.chart import BarChart,Reference

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
#cell = sheet['a1'] #method one to access cell object
cell = sheet.cell(1,1) #another method to access cell object

#print cell value:
#print(cell.value)

####ITERATE THROUGH WORKBOOK#### 

#1.need to know how many rows we have:
#print(sheet.max_row)

#2. add a for loop to generate #s 1 to 4.
for row in range(2,sheet.max_row + 1):
    cell = sheet.cell(row,3)
    corrected_price =cell.value * 0.9
    corrected_price_cell = sheet.cell(row,4) #adding a new collumn
    corrected_price_cell.value = corrected_price

#3. add a chart
#create an instance of the Reference class and store it in object
values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)

#create an instance of the Barchart class and store it in object
chart = BarChart()
chart.add_data(values) # chart is created in memory only at this point
sheet.add_chart(chart, "a6")
wb.save('transactions2.xlsx')
